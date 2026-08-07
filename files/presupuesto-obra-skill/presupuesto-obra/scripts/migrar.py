#!/usr/bin/env python3
"""
migrar.py — Sprint 0: depuracion de maestros y generacion de CSV para carga.

Produce en la carpeta de salida:
    01_sucursales.csv
    02_familias.csv
    03_articulos.csv          (con flag activo segun cobertura de precio)
    04_cuentas.csv            (maestro + cuentas huerfanas halladas en obras)
    05_precios.csv            (formato ancho: articulo x sucursal x 4 anios)
    99_diagnostico.json
    99_diagnostico.md

Uso:
  python migrar.py --articulos ZMAESTRO_DE_ARTICULOS.xls \
                   --cuentas   Nivel_10_Transformado.xlsx \
                   --precios   ZPRECIOS_MARVAL.xlsm \
                   --obras     BAIKAL_TORRE_3_V01_PTO_CON_APUs.xls [otra.xls ...] \
                   --out       migracion/
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import warnings
from collections import Counter, defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
warnings.filterwarnings("ignore")

from presupuesto_core import (clave_catalogo, nivel_de, normaliza_codigo,  # noqa
                             padre_de, plantilla_de, tipo_recurso_de, RE_ART)
import fuentes_marval as fm  # noqa

SUCURSALES = [
    ("BUC", "BUCARAMANGA"), ("BOG", "BOGOTA"), ("BAQ", "BARRANQUILLA"),
    ("CTG", "CARTAGENA"), ("CAL", "CALI"), ("ZIP", "ZIPAQUIRA"),
    ("RIC", "RICAURTE"),
]

# Nombres de familia conocidos. Completar con el area de costos.
FAMILIAS_CONOCIDAS = {
    "02": "CONTRATOS DE OBRA", "06": "ADITIVOS Y QUIMICOS",
    "07": "AGREGADOS", "15": "CEMENTOS Y CALES", "19": "CONCRETOS",
    "27": "ACEROS", "43": "TOPOGRAFIA",
}


def _w(path, header, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    return len(rows)


# --------------------------------------------------------------------------
# Matriz completa de precios (todas las sucursales, todos los anios)
# --------------------------------------------------------------------------

def matriz_precios(path: str) -> list[dict]:
    """Devuelve filas {articulo, sucursal, tipo, lista, anio_base, p1..p4}."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    filas: list[dict] = []

    def limpio(v):
        if v in (None, "", "#VALUE!", "#N/A", "#REF!", "#DIV/0!"):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    # ---- materiales: bloques de 5 columnas (lista + 4 anios) --------------
    if "PRECIOS MATERIALES" in wb.sheetnames:
        ws = wb["PRECIOS MATERIALES"]
        it = ws.iter_rows(values_only=True)
        head_suc = head_anio = None
        for i, r in enumerate(it):
            if i == 7:
                head_suc = list(r)
            elif i == 8:
                head_anio = list(r)
                break
        bloques = []
        for i, s in enumerate(head_suc or []):
            if s and str(s).strip().upper() in dict(SUCURSALES).values():
                if str(head_anio[i]).strip().upper().startswith("LISTA"):
                    anios = [head_anio[i + k] for k in range(1, 5)]
                    bloques.append((str(s).strip().upper(), i, anios))
        for r in ws.iter_rows(min_row=10, values_only=True):
            if not r or r[0] is None:
                continue
            cod = str(r[0]).strip()
            for suc, ci, anios in bloques:
                precios = [limpio(r[ci + k]) if ci + k < len(r) else None
                           for k in range(1, 5)]
                if not any(precios):
                    continue
                filas.append({
                    "articulo": cod, "sucursal": suc, "tipo": "MAT",
                    "lista": str(r[ci] or "").strip(),
                    "anio_base": int(str(anios[0]).strip()),
                    "p1": precios[0], "p2": precios[1],
                    "p3": precios[2], "p4": precios[3],
                    "alcance": "", "observacion": "",
                })

    # ---- contratos: bloques de 4 columnas (4 anios) ----------------------
    if "LISTAS DE PRECIOS MARVAL" in wb.sheetnames:
        ws = wb["LISTAS DE PRECIOS MARVAL"]
        cab = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        head_suc, head_anio = list(cab[3]), list(cab[5])
        bloques, i = [], 0
        while i < len(head_suc):
            s = head_suc[i]
            if s and str(s).strip().upper() in dict(SUCURSALES).values() and \
                    str(head_anio[i]).strip().isdigit():
                bloques.append((str(s).strip().upper(), i,
                                int(str(head_anio[i]).strip())))
                i += 4
            else:
                i += 1
        for r in ws.iter_rows(min_row=8, values_only=True):
            if not r or r[2] is None:
                continue
            cod = str(r[2]).strip()
            for suc, ci, a0 in bloques:
                precios = [limpio(r[ci + k]) if ci + k < len(r) else None
                           for k in range(4)]
                if not any(precios):
                    continue
                filas.append({
                    "articulo": cod, "sucursal": suc, "tipo": "CONTRATO",
                    "lista": "LISTA NACIONAL",
                    "anio_base": a0,
                    "p1": precios[0], "p2": precios[1],
                    "p3": precios[2], "p4": precios[3],
                    "alcance": str(r[38] or "")[:250] if len(r) > 38 else "",
                    "observacion": str(r[39] or "")[:250] if len(r) > 39 else "",
                })
    return filas


# --------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--articulos", required=True)
    p.add_argument("--cuentas", required=True)
    p.add_argument("--precios", required=True)
    p.add_argument("--obras", nargs="*", default=[])
    p.add_argument("--out", default="migracion")
    a = p.parse_args()
    os.makedirs(a.out, exist_ok=True)
    D: dict = {}

    print("Leyendo fuentes...")
    articulos = fm.cargar_articulos(a.articulos)
    cuentas = fm.cargar_cuentas(a.cuentas)
    precios = matriz_precios(a.precios)
    obras = [fm.importar_presupuesto(o) for o in a.obras]

    # ---- 1. sucursales ---------------------------------------------------
    n = _w(os.path.join(a.out, "01_sucursales.csv"),
           ["codigo", "nombre", "activa", "sucursal_referencia"],
           [[c, nm, "true", "BUCARAMANGA" if nm != "BUCARAMANGA" else ""]
            for c, nm in SUCURSALES])
    D["sucursales"] = n

    # ---- 2. familias -----------------------------------------------------
    fams = sorted({v["familia"] for v in articulos.values()
                   if v["familia"] and v["familia"].lower() != "nan"})
    uso = Counter(v["familia"] for v in articulos.values())
    n = _w(os.path.join(a.out, "02_familias.csv"),
           ["codigo", "nombre", "n_articulos", "revisar_nombre"],
           [[f, FAMILIAS_CONOCIDAS.get(f, f"FAMILIA {f}"), uso[f],
             "false" if f in FAMILIAS_CONOCIDAS else "true"] for f in fams])
    D["familias"] = {"total": n,
                     "sin_nombre": sum(1 for f in fams if f not in FAMILIAS_CONOCIDAS)}

    # ---- 3. articulos ----------------------------------------------------
    con_precio = defaultdict(set)
    for r in precios:
        con_precio[r["articulo"]].add(r["sucursal"])
    usados = Counter()
    for o in obras:
        for l in o.lineas:
            for i in l.insumos:
                usados[clave_catalogo(i.codigo)] += 1

    rows, inactivos, malformados = [], 0, []
    for cod, v in sorted(articulos.items()):
        sucs = con_precio.get(cod, set())
        activo = bool(sucs) or cod in usados
        if not activo:
            inactivos += 1
        rows.append([cod, v["descripcion"], v["um"], v["familia"],
                     v["tipo_linea"], "true" if activo else "false",
                     len(sucs), usados.get(cod, 0)])
    n = _w(os.path.join(a.out, "03_articulos.csv"),
           ["codigo", "descripcion", "unidad_medida", "familia", "tipo_linea",
            "activo", "n_sucursales_con_precio", "n_usos_en_obras"], rows)
    D["articulos"] = {"total": n, "inactivos_sugeridos": inactivos,
                      "activos": n - inactivos,
                      "pct_inactivos": round(inactivos / n * 100, 1)}

    # ---- 4. cuentas ------------------------------------------------------
    faltantes: dict[str, dict] = {}
    for o in obras:
        for l in o.lineas:
            if l.codigo not in cuentas and l.codigo not in faltantes:
                faltantes[l.codigo] = {"codigo": l.codigo, "nivel": l.nivel,
                                       "descripcion": l.descripcion,
                                       "um": l.unidad, "origen": o.proyecto}
    # completar ancestros faltantes
    for cod in list(faltantes):
        for anc in [padre_de(cod)]:
            while anc:
                if anc not in cuentas and anc not in faltantes:
                    faltantes[anc] = {"codigo": anc, "nivel": nivel_de(anc),
                                      "descripcion": "*** POR DEFINIR ***",
                                      "um": "UN", "origen": "derivado"}
                anc = padre_de(anc)

    rows = []
    for cod, v in sorted({**cuentas, **{}}.items()):
        rows.append([cod, v["nivel"], padre_de(cod) or "", v["descripcion"],
                     v.get("um", ""), plantilla_de(cod) or "ESPECIAL",
                     "true", "MAESTRO"])
    for cod, v in sorted(faltantes.items()):
        rows.append([cod, v["nivel"], padre_de(cod) or "", v["descripcion"],
                     v["um"], plantilla_de(cod) or "ESPECIAL", "true",
                     f"ALTA_REQUERIDA:{v['origen']}"])
    n = _w(os.path.join(a.out, "04_cuentas.csv"),
           ["codigo", "nivel", "codigo_padre", "descripcion", "unidad_medida",
            "plantilla", "activa", "origen"], rows)
    por_plantilla = Counter(plantilla_de(c) or "ESPECIAL"
                            for c in {**cuentas, **faltantes})
    D["cuentas"] = {"total": n, "del_maestro": len(cuentas),
                    "alta_requerida": len(faltantes),
                    "sin_descripcion": sum(1 for v in faltantes.values()
                                           if v["descripcion"].startswith("***")),
                    "por_plantilla": dict(por_plantilla)}

    # ---- 5. precios ------------------------------------------------------
    rows = [[r["articulo"], r["sucursal"], r["tipo"], r["lista"], r["anio_base"],
             r["p1"], r["p2"], r["p3"], r["p4"], r["alcance"], r["observacion"]]
            for r in precios]
    n = _w(os.path.join(a.out, "05_precios.csv"),
           ["articulo", "sucursal", "tipo", "lista_origen", "anio_base",
            "precio_anio_1", "precio_anio_2", "precio_anio_3", "precio_anio_4",
            "alcance", "observacion"], rows)
    cob = Counter(r["sucursal"] for r in precios)
    bases = defaultdict(set)
    for r in precios:
        bases[r["sucursal"]].add(r["anio_base"])
    D["precios"] = {"total": n,
                    "por_sucursal": dict(cob),
                    "anio_base_por_sucursal": {k: sorted(v) for k, v in bases.items()},
                    "articulos_con_precio": len(con_precio)}

    # ---- 6. hallazgos de obras ------------------------------------------
    mal = []
    for o in obras:
        for l in o.lineas:
            for i in l.insumos:
                c = normaliza_codigo(i.codigo)
                if not RE_ART.match(c):
                    mal.append({"obra": o.proyecto, "n10": l.codigo,
                                "codigo": c, "descripcion": i.descripcion})
                elif clave_catalogo(c) not in articulos:
                    mal.append({"obra": o.proyecto, "n10": l.codigo,
                                "codigo": c, "descripcion": i.descripcion})
    D["obras"] = {"analizadas": [o.proyecto for o in obras],
                  "insumos_problematicos": mal}

    # ---- diagnostico -----------------------------------------------------
    with open(os.path.join(a.out, "99_diagnostico.json"), "w", encoding="utf-8") as fh:
        json.dump(D, fh, ensure_ascii=False, indent=2)

    md = ["# Diagnostico Sprint 0 — depuracion de maestros", "",
          "Generado por `migrar.py` sobre los archivos fuente reales.", "",
          "## Resumen", "",
          "| Maestro | Registros | Accion requerida |", "|---|---|---|",
          f"| Sucursales | {D['sucursales']} | Confirmar cuales quedan activas |",
          f"| Familias | {D['familias']['total']} | **{D['familias']['sin_nombre']} sin nombre descriptivo** |",
          f"| Articulos | {D['articulos']['total']} | **{D['articulos']['inactivos_sugeridos']} sugeridos como inactivos ({D['articulos']['pct_inactivos']}%)** |",
          f"| Cuentas | {D['cuentas']['total']} | **{D['cuentas']['alta_requerida']} altas requeridas** |",
          f"| Precios | {D['precios']['total']} | Cobertura desigual por sucursal |",
          "",
          "## 1. Articulos", "",
          f"- Total en el maestro: **{D['articulos']['total']}**",
          f"- Con precio vigente en al menos una sucursal, o usados en obra: "
          f"**{D['articulos']['activos']}**",
          f"- Sin precio y sin uso -> marcar `activo = false`: "
          f"**{D['articulos']['inactivos_sugeridos']}** ({D['articulos']['pct_inactivos']}%)",
          "",
          "> No borrar: los articulos inactivos siguen referenciados por",
          "> presupuestos historicos. Solo se ocultan del buscador.",
          "",
          "## 2. Cuentas", "",
          f"- En `Nivel_10_Transformado.xlsx`: **{D['cuentas']['del_maestro']}**",
          f"- Usadas en obras reales pero ausentes del maestro: "
          f"**{D['cuentas']['alta_requerida']}**",
          f"- De esas, sin descripcion conocida (ancestros derivados): "
          f"**{D['cuentas']['sin_descripcion']}** — requieren definicion manual",
          "", "Distribucion por plantilla:", ""]
    md += [f"- `{k}`: {v}" for k, v in sorted(D["cuentas"]["por_plantilla"].items())]
    md += ["", "## 3. Precios", "",
           "| Sucursal | Registros | Anio base |", "|---|---|---|"]
    for s, c in sorted(D["precios"]["por_sucursal"].items(), key=lambda x: -x[1]):
        md.append(f"| {s} | {c} | {', '.join(map(str, D['precios']['anio_base_por_sucursal'][s]))} |")
    md += ["",
           f"Articulos distintos con al menos un precio: "
           f"**{D['precios']['articulos_con_precio']}** de {D['articulos']['total']} "
           f"({D['precios']['articulos_con_precio'] / D['articulos']['total'] * 100:.1f}%).",
           "",
           "> Los anios base difieren entre sucursales. Al cargar, respetar el",
           "> campo `anio_base` de cada fila: `precio_anio_1` NO es siempre 2025.",
           "", "## 4. Insumos problematicos en obras", ""]
    if mal:
        md += ["| Obra | Subactividad | Codigo | Descripcion |", "|---|---|---|---|"]
        md += [f"| {m['obra']} | {m['n10']} | `{m['codigo']}` | {m['descripcion']} |"
               for m in mal]
    else:
        md.append("Ninguno.")
    md += ["", "## 5. Orden de carga en Zoho Creator", "",
           "```", "01_sucursales.csv", "02_familias.csv",
           "03_articulos.csv    (depende de familias)",
           "04_cuentas.csv      (cargar por nivel: 4, luego 5, 8, 10)",
           "05_precios.csv      (depende de articulos y sucursales)", "```", "",
           "Cargar por Bulk API en lotes de 200, respetando el limite de 50",
           "llamadas por minuto.", ""]

    with open(os.path.join(a.out, "99_diagnostico.md"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(md))

    print(f"\n-> {a.out}/")
    for k in ("sucursales", "familias", "articulos", "cuentas", "precios"):
        v = D[k]
        print(f"   {k:<12} {v if isinstance(v, int) else v.get('total')}")
    print(f"   diagnostico  99_diagnostico.md / .json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
