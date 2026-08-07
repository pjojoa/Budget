"""
fuentes_marval.py — Lectores de los archivos fuente de Marval.

Cubre cuatro fuentes reales:
  ZMAESTRO_DE_ARTICULOS.xls        hoja 'MAESTRO ARTICULOS'
  Nivel_10_Transformado.xlsx       hoja 'Nivel 10'   (maestro de cuentas aplanado)
  ZPRECIOS_MARVAL.xlsm             hojas 'PRECIOS MATERIALES' y 'LISTAS DE PRECIOS MARVAL'
  <PROYECTO>_PTO_CON_APUs.xls      hoja 'PTO CON APUs' (presupuesto exportado)

Requiere: pandas, openpyxl, xlrd>=2.0 (solo para .xls).
"""

from __future__ import annotations

import re
import warnings
from decimal import Decimal

from presupuesto_core import (Insumo, Linea, Obra, _d, clave_catalogo,
                              nivel_de, normaliza_codigo, tipo_recurso_de)

warnings.filterwarnings("ignore")

# --------------------------------------------------------------------------
# Maestro de articulos
# --------------------------------------------------------------------------

def cargar_articulos(path: str) -> dict[str, dict]:
    """codigo -> {descripcion, um, familia, tipo_linea}"""
    import pandas as pd
    df = pd.read_excel(path, engine="xlrd" if path.lower().endswith(".xls") else None)
    df.columns = ["cod", "descripcion", "um", "familia", "tipo_linea"][: len(df.columns)]
    out = {}
    for r in df.itertuples(index=False):
        cod = str(r.cod).strip()
        if not cod or cod.lower() == "nan":
            continue
        out[cod] = {"codigo": cod, "descripcion": str(r.descripcion).strip(),
                    "um": str(r.um).strip(), "familia": str(r.familia).strip(),
                    "tipo_linea": str(r.tipo_linea).strip()}
    return out


# --------------------------------------------------------------------------
# Maestro de cuentas (Nivel 10 transformado)
# --------------------------------------------------------------------------

def cargar_cuentas(path: str) -> dict[str, dict]:
    """Reconstruye las cuentas N4/N5/N8/N10 a partir del archivo aplanado."""
    import pandas as pd
    df = pd.read_excel(path)
    out: dict[str, dict] = {}

    def put(cod, desc, nivel, um=""):
        cod = normaliza_codigo(cod)
        if cod and cod.lower() != "nan" and cod not in out:
            out[cod] = {"codigo": cod, "descripcion": str(desc).strip(),
                        "nivel": nivel, "um": str(um).strip()}

    for r in df.itertuples(index=False):
        d = r._asdict() if hasattr(r, "_asdict") else dict(zip(df.columns, r))
        vals = list(d.values())
        c10, c8, c5, c4 = [str(v) for v in vals[0:4]]
        d10, d8, d5, d4 = [str(v) for v in vals[4:8]]
        um = str(vals[12]) if len(vals) > 12 else ""
        put(c4, d4, 4)
        put(c5, d5, 5)
        put(c8, d8, 8)
        put(c10, d10, 10, um)
    return out


# --------------------------------------------------------------------------
# Catalogo de precios
# --------------------------------------------------------------------------

def _norm(s) -> str:
    s = str(s or "").strip().upper()
    for a, b in (("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U")):
        s = s.replace(a, b)
    return s


def cargar_precios(path: str, sucursal: str, anio: int) -> dict[str, dict]:
    """Devuelve clave_catalogo -> {precio, descripcion, um, origen, tipo}.

    Une la hoja de materiales (codigos numericos) con la lista de contratos
    (mano de obra / todo costo / equipo, referenciados con prefijo MO/TC/EQ).
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    suc = _norm(sucursal)
    out: dict[str, dict] = {}

    def col_anio(cabecera_suc, cabecera_anio, base_offset=0):
        """Localiza la columna cuya sucursal y anio coinciden."""
        for i, (s, a) in enumerate(zip(cabecera_suc, cabecera_anio)):
            if _norm(s) == suc and str(a).strip() == str(anio):
                return i + base_offset
        return None

    # ---- materiales -------------------------------------------------------
    if "PRECIOS MATERIALES" in wb.sheetnames:
        ws = wb["PRECIOS MATERIALES"]
        rows = ws.iter_rows(values_only=True)
        head_suc = head_anio = None
        for i, r in enumerate(rows):
            if i == 7:
                head_suc = list(r)
            if i == 8:
                head_anio = list(r)
                break
        ci = col_anio(head_suc, head_anio) if head_suc else None
        if ci is not None:
            for r in ws.iter_rows(min_row=10, values_only=True):
                if not r or r[0] is None:
                    continue
                cod = str(r[0]).strip()
                val = r[ci] if ci < len(r) else None
                if val in (None, "", "#VALUE!", "#N/A"):
                    continue
                out[cod] = {"codigo": cod, "descripcion": str(r[1] or "").strip(),
                            "um": str(r[2] or "").strip(),
                            "familia": str(r[3] or "").strip(),
                            "precio": _d(val), "tipo": "MAT",
                            "origen": str(r[ci - (ci - 5) % 5] or "")[:40]}

    # ---- contratos (MO / TC / EQ) ----------------------------------------
    if "LISTAS DE PRECIOS MARVAL" in wb.sheetnames:
        ws = wb["LISTAS DE PRECIOS MARVAL"]
        rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        head_suc, head_anio = list(rows[3]), list(rows[5])
        ci = col_anio(head_suc, head_anio)
        if ci is not None:
            for r in ws.iter_rows(min_row=8, values_only=True):
                if not r or r[2] is None:
                    continue
                cod = str(r[2]).strip()
                val = r[ci] if ci < len(r) else None
                if val in (None, "", "#VALUE!", "#N/A"):
                    continue
                out.setdefault(cod, {})
                out[cod].update({"codigo": cod,
                                 "descripcion": str(r[3] or "").strip(),
                                 "um": str(r[4] or "").strip(),
                                 "familia": str(r[5] or "").strip(),
                                 "precio": _d(val), "tipo": "CONTRATO",
                                 "alcance": str(r[38] or "")[:200]})
    return out


# --------------------------------------------------------------------------
# Presupuesto exportado (PTO CON APUs)
# --------------------------------------------------------------------------

CAB = ["cod", "nd", "desc", "tipo", "un", "cant", "vu", "vt", "vinm", "vm2", "pct"]

TIPOS_APU = {"M.O", "T.C", "MAT", "EQ"}


def _tipo_apu(v) -> str:
    """El origen trae valores sucios ('OK', vacios). Normaliza al enum valido."""
    t = str(v or "").strip().upper()
    t = {"MO": "M.O", "TC": "T.C", "M O": "M.O"}.get(t, t)
    return t if t in TIPOS_APU else ""


def importar_presupuesto(path: str, hoja: str = "PTO CON APUs") -> Obra:
    """Lee el formato FT-PGC-PLA-002 y devuelve una Obra recalculada."""
    import pandas as pd
    eng = "xlrd" if path.lower().endswith(".xls") else None
    raw = pd.read_excel(path, engine=eng, sheet_name=hoja, header=None)

    def celda(f, c):
        try:
            v = raw.iat[f, c]
            return "" if pd.isna(v) else v
        except Exception:
            return ""

    obra = Obra(
        proyecto=str(celda(6, 3)).strip(),
        version=str(celda(9, 2)).strip(),
        n_inmuebles=int(_d(celda(11, 2) or 1)),
        area_inmueble_m2=_d(celda(12, 2) or 1),
    )

    # localizar la fila de encabezados de la tabla
    fila_h = 16
    for i in range(raw.shape[0]):
        if str(celda(i, 0)).strip().upper().startswith("CÓDIGO COSTO") or \
           str(celda(i, 0)).strip().upper().startswith("CODIGO COSTO"):
            fila_h = i
            break

    df = raw.iloc[fila_h + 1:, : len(CAB)].copy()
    df.columns = CAB

    actual10: Linea | None = None
    for r in df.itertuples(index=False):
        cod = str(r.cod).strip()
        if not cod or cod.lower() == "nan":
            continue
        nd = str(r.nd).strip().replace(".0", "")
        if nd not in ("4", "5", "8", "10", "11"):
            continue
        desc = "" if str(r.desc) == "nan" else str(r.desc).strip()
        un = "" if str(r.un) == "nan" else str(r.un).strip()
        cant = _d(0 if str(r.cant) == "nan" else r.cant)

        if nd == "11":
            if actual10 is None:
                continue
            actual10.insumos.append(Insumo(
                codigo=normaliza_codigo(cod), descripcion=desc, unidad=un,
                rendimiento=cant, precio=_d(0 if str(r.vu) == "nan" else r.vu),
                tipo=tipo_recurso_de(cod)))
            continue

        linea = Linea(codigo=normaliza_codigo(cod), nivel=int(nd), descripcion=desc,
                      unidad=un or "UN", cantidad=cant, tipo=_tipo_apu(r.tipo))
        obra.lineas.append(linea)
        actual10 = linea if nd == "10" else actual10

    return obra.recalcular()


# --------------------------------------------------------------------------
# Repricing
# --------------------------------------------------------------------------

def reprecio(obra: Obra, precios: dict[str, dict],
             sucursal: str = "", anio: int = 0) -> dict:
    """Sustituye los precios de todos los insumos por los del catalogo dado."""
    cambiados = no_encontrados = 0
    detalle = []
    for l in obra.lineas:
        if l.nivel != 10:
            continue
        for ins in l.insumos:
            k = clave_catalogo(ins.codigo)
            reg = precios.get(k)
            if not reg or not reg.get("precio"):
                no_encontrados += 1
                detalle.append({"codigo": ins.codigo, "estado": "SIN_PRECIO"})
                continue
            nuevo = _d(reg["precio"])
            if nuevo != _d(ins.precio):
                detalle.append({"codigo": ins.codigo, "anterior": float(ins.precio),
                                "nuevo": float(nuevo),
                                "var_pct": float((nuevo - _d(ins.precio)) /
                                                 (_d(ins.precio) or Decimal(1)) * 100)})
                cambiados += 1
            ins.precio = nuevo
    anterior = obra.total
    obra.recalcular()
    if sucursal:
        obra.sucursal = sucursal
    if anio:
        obra.anio_precios = anio
    return {"insumos_actualizados": cambiados, "sin_precio": no_encontrados,
            "total_anterior": float(anterior), "total_nuevo": float(obra.total),
            "variacion_pct": float((obra.total - anterior) / (anterior or Decimal(1)) * 100),
            "detalle": detalle}
